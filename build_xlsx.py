"""Produce a Pipedrive-ready XLSX with two sheets: Organizations + People.

Pipedrive import conventions followed:
- Organizations: Name (required), Label, Address, Website, Phone, Notes.
- People: Name (required), Email, Phone, Organization, Job Title, Label, Notes.

Extra columns (Priorité, Catégorie_rôle, Aire_thérapeutique, Séniorité, Source,
Flag_C_level, Flag_manager_de_managers, Décideur_ou_influenceur, Priority_score)
are kept as custom fields — Nicolas can map them at import.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ORG_COLUMNS = [
    "Name",
    "Website",
    "Label",
    "Address",
    "Phone",
    "Priorité (P1/P2/P3)",
    "Statut CRM",
    "Work Status",
    "PIC muchbetter",
    "Next Step",
    "CRM_ID",
    "LinkedIn URL",
    "Effectif",
    "Aires thérapeutiques (enrichies)",
    "Nb contacts mappés",
    "Date enrichissement",
    "Commentaires CRM",
    "Notes",
]

PEOPLE_COLUMNS = [
    "Name",
    "Organization",
    "Job Title",
    "Email",
    "Phone",
    "Label",
    "LinkedIn URL",
    "Location",
    "Priorité compte",
    "CRM_ID compte",
    "Séniorité",
    "Catégorie rôle",
    "Aire thérapeutique",
    "Flag C-Level",
    "Flag Manager-de-managers",
    "Flag BU Head",
    "Priority Score",
    "Source",
    "Déjà connu Nicolas (O/N)",
    "Décideur ou Influenceur (à remplir terrain)",
    "Notes",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autofit(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), max_width))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, max_width))


def write_workbook(
    path: Path,
    orgs: Iterable[dict],
    people: Iterable[dict],
) -> None:
    wb = Workbook()

    ws_org = wb.active
    ws_org.title = "Organizations"
    ws_org.append(ORG_COLUMNS)
    for row in orgs:
        ws_org.append([row.get(col, "") for col in ORG_COLUMNS])
    _style_header(ws_org, len(ORG_COLUMNS))
    _autofit(ws_org)

    ws_people = wb.create_sheet("People")
    ws_people.append(PEOPLE_COLUMNS)
    for row in people:
        ws_people.append([row.get(col, "") for col in PEOPLE_COLUMNS])
    _style_header(ws_people, len(PEOPLE_COLUMNS))
    _autofit(ws_people)

    # Legend sheet: documents the label conventions for Nicolas
    ws_legend = wb.create_sheet("Légende")
    legend_rows = [
        ["Champ", "Valeurs possibles", "Usage"],
        ["Label (Org)", "P1+ / P1 / P2 / P3", "Priorité commerciale du compte."],
        ["Label (People)", "C-Level, Digital, Data/AI, IT/IS, Medical Affairs, "
                            "Market Access, Commercial Excellence, R&D/Clinical, "
                            "HR/People, Marketing/Brand, Regulatory/Quality",
         "Catégorie principale — multi-label OK, séparé par virgule."],
        ["Séniorité", "Top (C-Level/GM) / VP / Head-Director / Senior Director / "
                      "Manager / IC / Junior", "Niveau hiérarchique estimé depuis le titre."],
        ["Priority Score", "0-100", "Score d'intérêt pour muchbetter.ai — "
                                     "Digital/Data/AI + C-Level + BU head > autres."],
        ["Flag C-Level", "Oui / Non", "Chief X / CEO / President / Country Manager / DG."],
        ["Flag Manager-de-managers", "Oui / Non (estimation)",
         "Heuristique: VP+ ou Senior Director+ ou Head of (global/EMEA). "
         "À valider sur le terrain."],
        ["Flag BU Head", "Oui / Non",
         "Responsable d'une Business Unit thérapeutique (oncologie, obésité, etc.)."],
        ["Source", "techtomed / mcp_enrich / crm",
         "Provenance du contact. techtomed = cité dans le CSV source."],
        ["Déjà connu Nicolas", "O / N",
         "O = déjà dans son CRM ou cité par TechToMed."],
        ["Décideur ou Influenceur", "(à remplir terrain)",
         "Volontairement vide — se remplit en conversation."],
    ]
    for r in legend_rows:
        ws_legend.append(r)
    _style_header(ws_legend, 3)
    _autofit(ws_legend, max_width=80)

    wb.save(path)


if __name__ == "__main__":
    # Smoke test
    out = Path("out") / f"mapping_{dt.date.today().isoformat()}_smoke.xlsx"
    out.parent.mkdir(exist_ok=True)
    orgs = [
        {
            "Name": "Novo Nordisk",
            "Website": "novonordisk.com",
            "Label": "P2",
            "Priorité (P1/P2/P3)": "P2",
            "CRM_ID": "2255",
            "PIC muchbetter": "Charles",
            "Statut CRM": "/",
            "Aires thérapeutiques (enrichies)": "Diabetes / Obesity",
            "Nb contacts mappés": 2,
            "Date enrichissement": dt.date.today().isoformat(),
            "Commentaires CRM": "Pauline Derrien, Nicolas Christol (patron BU obésité)",
        }
    ]
    people = [
        {
            "Name": "Nicolas Christol",
            "Organization": "Novo Nordisk",
            "Job Title": "Head of BU Obésité",
            "Label": "BU Head",
            "Priorité compte": "P2",
            "CRM_ID compte": "2255",
            "Séniorité": "Head / Director",
            "Catégorie rôle": "",
            "Aire thérapeutique": "Diabetes / Obesity / Metabolic",
            "Flag C-Level": "Non",
            "Flag Manager-de-managers": "Non",
            "Flag BU Head": "Oui",
            "Priority Score": 25,
            "Source": "techtomed",
            "Déjà connu Nicolas (O/N)": "O",
        }
    ]
    write_workbook(out, orgs, people)
    print(f"wrote {out}")
